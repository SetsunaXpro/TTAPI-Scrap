import os
from urllib.parse import urlparse
import asyncio
import aiohttp
import time
import pandas as pd
import datetime
from TikTokApi import TikTokApi
from TikTokApi.exceptions import EmptyResponseException, CaptchaException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def is_tiktok_url(url: str) -> bool:
    """Check if URL belongs to TikTok domains"""
    parsed = urlparse(url)
    return parsed.netloc.endswith('tiktok.com')

def is_tiktok_content_url(url: str) -> bool:
    """Check if URL is a valid TikTok video/photo post"""
    parsed = urlparse(url)
    if not parsed.netloc.endswith('tiktok.com'):
        return False
    
    path = parsed.path.lower()
    return any(x in path for x in ['/video/', '/photo/']) or '@' in path and ('/video' in path or '/photo' in path)



async def get_tiktok_data(video_url: str, api: TikTokApi) -> dict:
    try:
        # Resolve short URLs
        if "vt.tiktok.com" in video_url:
            resolved_url = await resolve_short_url(video_url)
            video_url = resolved_url

        if not is_tiktok_content_url(video_url):
            return {'status': 'error', 'message': 'Not a content URL'}

        modified_url = video_url.replace("/photo/", "/video/")

        video = api.video(url=modified_url)
        data = await video.info()
        
        is_photo_post = 'imagePost' in data
        stats = data.get("stats", {})
        views = stats.get("playCount") or stats.get("viewCount") or 0
        post_id = data.get("id")
        
        create_time = None
        create_time_ts = data.get("createTime")
        if create_time_ts:
            try:
                utc_time = datetime.datetime.fromtimestamp(
                    int(create_time_ts),
                    tz=datetime.timezone.utc
                )
                create_time = utc_to_local(utc_time)
            except Exception as e:
                print(f"Time error: {str(e)}")
                create_time = None

        return {
            'views': views,
            'create_time': create_time,
            'post_id': str(post_id) if post_id else None,
            'is_photo': is_photo_post
        }
    except Exception as e:
        return {'status': 'error', 'message': str(e)}

def utc_to_local(utc_dt):
    """Convert UTC datetime to local naive datetime"""
    return utc_dt.replace(tzinfo=datetime.timezone.utc).astimezone(tz=None).replace(tzinfo=None)

async def resolve_short_url(url: str) -> str:
    """Resolve shortened TikTok URLs to their canonical form"""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.head(
                url, 
                allow_redirects=True, 
                timeout=aiohttp.ClientTimeout(total=10)
            ) as response:
                return str(response.url)
    except Exception as e:
        print(f"URL resolution error: {str(e)}")
        return url  

async def process_urls(urls: list, api: TikTokApi) -> tuple:
    results = []
    seen_ids = set()
    for index, url in enumerate(urls):
        print(f"Processing ({index+1}/{len(urls)}): {url}")
        try:
            data = await get_tiktok_data(url, api)
            
            if 'status' in data:
                results.append(data)
                continue
                
            if data['post_id']:
                if data['post_id'] in seen_ids:
                    results.append({'status': 'duplicate'})
                else:
                    seen_ids.add(data['post_id'])
                    results.append(data)
            else:
                results.append({'status': 'error', 'message': 'No post ID found'})
                
        except Exception as e:
            print(f"Fatal error: {str(e)}")
            results.append({'status': 'error', 'message': str(e)})
        await asyncio.sleep(0.1)
    return results, seen_ids

async def main():
    try:
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # ARGB format
        skipped_indices = []
        
        try:
            wb_existing = load_workbook("urls.xlsx")
            ws_existing = wb_existing.active
            for excel_row in ws_existing.iter_rows(min_row=2):
                cell_f = excel_row[5]
                
                if cell_f.fill:
                    # Get color in ARGB format
                    cell_color = cell_f.fill.start_color
                    if isinstance(cell_color.rgb, str):
                        # Compare first 6 characters ignoring alpha channel
                        is_red = cell_color.rgb[-6:].lower() == 'ff0000'
                    else:
                        is_red = cell_color.index == 2
                    
                    if is_red:
                        df_index = excel_row[0].row - 2
                        skipped_indices.append(df_index)
                        print(f"Found red cell at Excel row {excel_row[0].row} (DF index {df_index})")
            
            wb_existing.close()
        except FileNotFoundError:
            pass

        # Now load the DataFrame normally
        df = pd.read_excel("urls.xlsx", dtype={9: str})
        print(f"Total DataFrame rows: {len(df)}")
        print(f"Skipping indices: {skipped_indices}")
        URL_COL = 8
        VIEWS_COL = 10
        # change date range
        DATE_RANGE_START = datetime.datetime(2025, 1, 23)
        DATE_RANGE_END = datetime.datetime(2025, 2, 23)

        if df.shape[1] <= 9:
            df.insert(9, 'Status', "")
        df.iloc[:, 9] = df.iloc[:, 9].fillna("").astype(str)

        valid_entries = []
        for idx, row in df.iterrows():
            if idx in skipped_indices:
                continue
                
            url = row.iloc[URL_COL]
            if pd.notna(url) and (url := str(url).strip()):
                # Skip non-TikTok URLs
                if not is_tiktok_url(url):
                    print(f"Skipping non-TikTok URL: {url}")
                    continue
                valid_entries.append((idx, url))

        urls = [url for _, url in valid_entries]

    except FileNotFoundError:
        print("Create urls.xlsx first!")
        return

    if not urls:
        print("No valid URLs found")
        return

    print(f"Starting scrape of {len(urls)} URLs...")
    start = time.time()

    async with TikTokApi() as api:
        await api.create_sessions(
            num_sessions=1,
            headless=True,
            ms_tokens=[os.getenv("MS_TOKEN")],
            sleep_after=3
        )
        results, seen_ids = await process_urls(urls, api)

    # Update DataFrame with results
    for i, res in enumerate(results):
        row_idx = valid_entries[i][0]
        status = res.get('status')
        message = res.get('message', '')
        
        # Handle special cases first
        if status == 'duplicate':
            df.iloc[row_idx, 9] = "Double submission"
            continue  # Skip to next iteration
        
        if status == 'error':
            if any(x in message for x in ['Not a content URL', 'Profile link']):
                df.iloc[row_idx, 9] = "Not a content submission"
            else:
                df.iloc[row_idx, 9] = f"Error: {message}"
            continue  # Skip to next iteration
        
        # Only process successful results below
        try:
            create_time = res.get('create_time')
            if create_time:
                if DATE_RANGE_START <= create_time <= DATE_RANGE_END:
                    df.iloc[row_idx, VIEWS_COL] = res.get('views', 0)
                    if not res.get('is_photo'):
                        df.iloc[row_idx, 9] = ""
                else:
                    df.iloc[row_idx, 9] = "Out of date submission"
            else:
                df.iloc[row_idx, 9] = "Missing creation time"
        except Exception as e:
            df.iloc[row_idx, 9] = f"Processing error: {str(e)}"

    # Save and format Excel
    with pd.ExcelWriter("urls.xlsx", engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    wb = load_workbook("urls.xlsx")
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        status = row[9].value
        if status and any(x in status for x in ["Double submission", "Out of date", "Error", "Not a content submission"]):
            for cell in row[0:9]:
                cell.fill = red_fill

    wb.save("urls.xlsx")
    print(f"Processed {len(results)} entries with {len(seen_ids)} unique posts")
    print(f"Completed in {time.time()-start:.2f} seconds")

if __name__ == "__main__":
    asyncio.run(main())
